import openpyxl 
import os.path 
import csv 
 
def generate_marksheet(nameDict, sub_name, LTPDict, student_info,  scoreDict, roll_no): 
    wb = openpyxl.Workbook() 
    sheet = wb.active 
    sheet.title="Overall"          
                                
    
    first_line=['Sl No.','Subject No.','Subject Name','L-T-P','Credit','Subject Type','Grade'] 
    prev_sem=0       
    ser_num=1           
    id_no=1          
     
    for row in student_info: 
        sheetname="Sem"+row[1]     
        if int(row[1])!=prev_sem:   
            ser_num=1                 
            prev_sem=int(row[1])    
            wb.create_sheet(index=id_no, title=sheetname)  
            wb[sheetname].append(first_line)                     
            id_no=id_no+1                             
 
         
        grade=[ser_num, row[2], sub_name[row[2]], LTPDict[row[2]], row[3], row[5], row[4]] 
        wb[sheetname].append(grade) 
        ser_num=ser_num+1                 
 
   
    sheet['A1']="Roll No."           
    sheet['B1']=roll_no           
    sheet['A2']="Name of Student"    
    sheet.merge_cells('B2:C2')      
    sheet.cell(row = 2, column = 2).value = nameDict[roll_no]   
    sheet['A3']="Discipline"         
    Branch=roll_no[4:6]       
    sheet['B3']=Branch          
    sheet['A4']="Semester No."       
    sheet['A5']="Semester wise Credit Taken"     
    sheet['A6']="SPI"                            
    sheet['A7']="Total Credits Taken"            
    sheet['A8']="CPI" 
 
    names_of_sheet=wb.sheetnames                  
    totalCred=0                            
    totalObtained=0  
 
    for i in range(1,len(names_of_sheet)): 
        sem_sheet=wb[names_of_sheet[i]]           
        sem_no=names_of_sheet[i][3:]              
        column_letter=openpyxl.utils.get_column_letter(i+1)  
        sheet[column_letter+str(4)]=int(sem_no)             
        semwise_credit=0                       
        semwise_obtained=0                     
        row_count = sem_sheet.max_row        
 
        for j in range(2, row_count + 1):  
            credit=int(sem_sheet['E'+str(j)].value)  
            grade=sem_sheet['G'+str(j)].value        
            semwise_credit=semwise_credit+credit     
            semwise_obtained=semwise_obtained+(credit*scoreDict[grade])  
        totalCred=totalCred+semwise_credit     
        totalObtained=totalObtained+semwise_obtained 
        SPI_obtained=semwise_obtained/semwise_credit          
        SPI_obtained=round(SPI_obtained,2)                   
        CPI_obtained=totalObtained/totalCred              
        CPI_obtained=round(CPI_obtained,2)                     
        sheet[column_letter+str(5)]=semwise_credit   
        sheet[column_letter+str(6)]=SPI_obtained              
        sheet[column_letter+str(7)]=totalCred     
        sheet[column_letter+str(8)]=CPI_obtained              
 
    wb.save(r'output\\'+ roll_no + '.xlsx')          
    return 
 
 
nameDict={}        
LTPDict={}          
sub_name={}      
 
with open('names-roll.csv') as f:     
    reader = csv.reader(f) 
    i=1 
    for line in reader: 
        if i==1:                      
            i=0 
            continue 
        nameDict[line[0]]=line[1]        
 
with open('subjects_master.csv') as f: 
    reader = csv.reader(f) 
    i=1 
    for line in reader: 
        if i==1:                         
            i=0 
            continue 
        sub_name[line[0]]=line[1]         
        LTPDict[line[0]]=line[2]        
 
try:
    os.mkdir('output')
except:
    pass

scoreDict={'F*':0, 'DD*':4, 'I':0, 'F':0, 'DD':4, 'CD':5, 'CC':6, 'BC':7, ' BB':8, 'BB':8, 'AB':9, 'AA':10 } 
 
roll_no="0401CS01"                
student_info=[]                   
file=open("grades.csv", "r")        
head = file.readline()              
 
for row in file:        
    row=row.split(',')          
    if row[0]==roll_no:         
        student_info.append(row) 
    else: 
        
        generate_marksheet(nameDict, sub_name, LTPDict, student_info,  scoreDict, roll_no) 
        student_info=[]     
        student_info.append(row) 
        roll_no=row[0] 
 

generate_marksheet(nameDict, sub_name, LTPDict, student_info, scoreDict, roll_no)